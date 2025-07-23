from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import traceback
import pandas as pd
import time
import openpyxl
import requests


def get_usd_to_inr_rate(url):
    
    response = requests.get(url)
    data = response.json()
    
    if not data.get('success', True):
        raise Exception(f"API Error: {data.get('error', {}).get('info', 'No info')}")
    
    rates = data.get('rates', {})
    
    # Fixer.io's base currency is EUR by default, so convert USD to INR via EUR rates
    usd_rate = rates.get('USD')
    inr_rate = rates.get('INR')
    
    if usd_rate is None or inr_rate is None:
        raise Exception("USD or INR rate missing in API response.")
    
    # Convert USD to INR: INR_per_USD = INR_per_EUR / USD_per_EUR
    usd_to_inr = inr_rate / usd_rate
    
    return round(usd_to_inr, 4)

# Extract scraped rates into pandas DataFrame
def extract_rates_to_df(html_source, hsn_code):
    soup = BeautifulSoup(html_source, "html.parser")

    rate_ids = {
        "Basic Customs Duty (BCD)": "t_bcd_rate",
        "Social Welfare Surcharge (SWC)": "t_scd_rate",
        "IGST Levy": "t_igst_rate",
    }

    data = {}
    for rate_name, span_id in rate_ids.items():
        span = soup.find("span", id=span_id)
        data[rate_name] = span.text.strip() if span and span.text.strip() else None

    data["HSN Code"] = hsn_code
    df = pd.DataFrame([data])
    return df

# Wait until certain element's text is non-empty
def wait_for_value(driver, by, element_id, timeout=30):
    wait = WebDriverWait(driver, timeout)
    wait.until(lambda d: d.find_element(by, element_id).text.strip() != "")

# Simple calculation of Total matching your Excel formula logic
def calculate_total(usd_to_inr, bcd_percent, swc_percent, igst_percent, c4_value, b5_value):
    bcd = float(bcd_percent or 0) / 100
    swc = float(swc_percent or 0) / 100
    igst = float(igst_percent or 0) / 100

    # swc = swc_percent of bcd 

    total_rate = bcd + swc + igst
    total = (c4_value * usd_to_inr) * total_rate + b5_value
    return total

# Create a new Excel file writing in values and the computed Total (no formula changes)
def create_new_excel(
        template_path, output_path,
        hsn_code,
        usd_inr, bcd, swc, igst,
        c4_val, b5_val
    ):
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Fill required input cells
    ws["C1"] = usd_inr
    ws["B10"] = bcd
    ws["B11"] = swc
    ws["B13"] = igst
    ws["C4"] = c4_val
    ws["B5"] = b5_val

    # # Put calculated Total in C20 (adjust cell if needed)
    # ws["C20"] = total_val

    wb.save(output_path)
    print(f"Saved output Excel: {output_path}")

def scrape_hsn_duty(hsn_code, c4_value, b5_value, usd_to_inr_rate, input_excel_path):
    chrome_options = Options()
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--start-maximized")
    # chrome_options.add_argument("--headless")  # uncomment to run headless

    driver = webdriver.Chrome(options=chrome_options)
    wait = WebDriverWait(driver, 25)

    try:
        # Load page and enter HSN
        driver.get("https://www.old.icegate.gov.in/Webappl/Trade-Guide-on-Imports")

        input_box = wait.until(EC.presence_of_element_located((By.NAME, "cth")))
        input_box.clear()
        input_box.send_keys(hsn_code)
        driver.find_element(By.ID, "submitbutton").click()

        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div#tmptest1")))

        # Find and click HSN row, scroll if needed
        def find_target_row():
            divs = driver.find_elements(By.CSS_SELECTOR, "div.row.rowh")
            for d in divs:
                if d.get_attribute("value") == hsn_code:
                    return d
            return None

        scroll_container = driver.find_element(By.CSS_SELECTOR, "div#tmptest1")

        target_row = find_target_row()
        if not target_row:
            last_height = -1
            for _ in range(15):
                driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", scroll_container)
                time.sleep(1)
                new_height = driver.execute_script("return arguments[0].scrollHeight", scroll_container)
                if new_height == last_height:
                    break
                last_height = new_height
                target_row = find_target_row()
                if target_row:
                    break

        if not target_row:
            print(f"HSN code {hsn_code} not found in tariff detail list.")
            return

        driver.execute_script("arguments[0].scrollIntoView(true);", target_row)
        time.sleep(0.3)

        try:
            wait.until(EC.element_to_be_clickable(target_row))
            target_row.click()
        except:
            driver.execute_script("arguments[0].click();", target_row)

        # Wait for AJAX updated rates to appear
        for eid in ["t_bcd_rate", "t_scd_rate", "t_igst_rate"]:
            try:
                wait_for_value(driver, By.ID, eid, 30)
            except:
                pass  # ignore if missing

        html_source = driver.page_source
        df_rates = extract_rates_to_df(html_source, hsn_code)

        print("\nExtracted Duty Rates:")
        print(df_rates)

        # Use scraped rates and input values to calculate Total
        bcd_val = df_rates.at[0, "Basic Customs Duty (BCD)"]
        swc_val = df_rates.at[0, "Social Welfare Surcharge (SWC)"]
        igst_val = df_rates.at[0, "IGST Levy"]

        # # Calculate in python
        # total = calculate_total(
        #     usd_to_inr=usd_to_inr_rate,
        #     bcd_percent=bcd_val,
        #     swc_percent=swc_val,
        #     igst_percent=igst_val,
        #     c4_value=c4_value,
        #     b5_value=b5_value
        # )

        # print(f"\nCalculated Total (Python): {total}")

        # Create new Excel file with input values and computed total
        output_excel_path = f"D:\\Github\\AgenticAI\\LangChain\\HSN\\data\\HSN_OUTPUT_FILES\\HSN_{hsn_code}_output.xlsx"
        create_new_excel(
            input_excel_path,
            output_excel_path,
            hsn_code,
            usd_to_inr_rate,
            bcd_val+"%",
            swc_val+"%",
            igst_val+"%",
            c4_value,
            b5_value
        )

    except Exception:
        print("An error occurred:")
        traceback.print_exc()

    finally:
        print("\nScript finished. Browser will stay open for inspection.")
        input("Press Enter to exit and close the browser...")
        driver.quit()


if __name__ == "__main__":
    # Example inputs:
    hsn_code = "73182100"            # Your HSN code
    c4_input = 1228                  # Value to put in cell C4
    b5_input = 0.095                   # Value to put in cell B5
    api_key = "564c4720a0cb45eda83bd8a481b98624"
    url = f"http://data.fixer.io/api/latest?access_key={api_key}"
    usd_inr_input = get_usd_to_inr_rate(url)          # USD to INR rate to put in C1
    usd_inr_input += 1.5
    print(usd_inr_input)
    input_excel_file = "D:\\Github\\AgenticAI\\LangChain\\HSN\\data\\Import Calculator With GST revised with BCD.xlsx"  # Path to your Excel template/reference

    scrape_hsn_duty(hsn_code, c4_input, b5_input, usd_inr_input, input_excel_file)
